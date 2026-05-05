#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║   🌈  N E T W O R K   N E R D A T H R O N   3 0 0 1 ™  🌈         ║
║        "Keep on Truckin'... But Ask Permission First!"               ║
║        The Grooviest Firewall in the Known Universe                  ║
║                                                                      ║
║  Changes vs 3000:                                                    ║
║   • Scope dropdown no longer hovers over Deny                        ║
║   • Adjustable splitter between connections & rules/log/stats        ║
║   • Duration: removed 'once', added '3 Hours' & 'Until logout'      ║
║   • Scope: 4th option 'This Port, Any IP'                           ║
║   • Rules panel shows live countdown to expiry                       ║
║   • Massively expanded port directory                                ║
║   • Export log to .xlsx                                              ║
║   • Enter key triggers Allow (green = default)                       ║
║   • All columns are drag-reorderable                                 ║
║   • Restart handled by restart_nerdathron.sh                        ║
╚══════════════════════════════════════════════════════════════════════╝

Requirements:
    pip install PyQt6 psutil openpyxl

Run (monitoring only):
    python3 nerdathron3001.py

Run (with connection blocking - recommended):
    sudo python3 nerdathron3001.py

Data is stored in: ~/.nerdathron3000/traffic.db
"""

import sys
import os
import json
import time
import socket
import sqlite3
import threading
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from dataclasses import dataclass
from typing import Optional, List

import psutil

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QTabWidget,
    QDialog, QComboBox, QLineEdit, QTextEdit, QSplitter,
    QSystemTrayIcon, QMenu, QFrame, QHeaderView, QScrollArea,
    QGroupBox, QSpinBox, QMessageBox, QAbstractItemView,
    QCheckBox, QProgressBar, QFileDialog
)
from PyQt6.QtCore import (
    Qt, QTimer, pyqtSignal, QThread, QSize, QRect, QPoint
)
from PyQt6.QtGui import (
    QColor, QFont, QPixmap, QIcon, QPainter, QLinearGradient,
    QBrush, QPen, QRadialGradient, QAction, QCursor
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


# ══════════════════════════════════════════════════════════════════════
#  🎨  THE GROOVY PALETTE — 70s Psychedelic Neon  (untouched)
# ══════════════════════════════════════════════════════════════════════

C = {
    'bg':        '#06000F',
    'panel':     '#0F0024',
    'card':      '#180038',
    'border':    '#3B1570',
    'purple':    '#BF5FFF',
    'violet':    '#7200CA',
    'pink':      '#FF2D78',
    'magenta':   '#FF00CC',
    'orange':    '#FF7700',
    'yellow':    '#FFE000',
    'gold':      '#FFB300',
    'lime':      '#39FF14',
    'green':     '#00CC55',
    'cyan':      '#00F5FF',
    'teal':      '#00C8B0',
    'blue':      '#4488FF',
    'white':     '#F5EEFF',
    'dim':       '#8866BB',
    'allow':     '#00FF88',
    'deny':      '#FF3333',
    'warn':      '#FFB800',
    'pending':   '#FF99FF',
}

# ══════════════════════════════════════════════════════════════════════
#  🔌  EXPANDED PORT DIRECTORY (was ~25 ports, now 90+)
# ══════════════════════════════════════════════════════════════════════

KNOWN_PORTS = {
    # Core internet
    20:  'FTP-Data',    21:  'FTP',         22:  'SSH',
    23:  'Telnet',      25:  'SMTP',        53:  'DNS',
    67:  'DHCP-S',      68:  'DHCP-C',      69:  'TFTP',
    80:  'HTTP',        88:  'Kerberos',    110: 'POP3',
    119: 'NNTP',        123: 'NTP',         135: 'RPC',
    137: 'NetBIOS',     138: 'NetBIOS',     139: 'NetBIOS',
    143: 'IMAP',        161: 'SNMP',        162: 'SNMP-Trap',
    179: 'BGP',         194: 'IRC',         389: 'LDAP',
    443: 'HTTPS',       445: 'SMB',         465: 'SMTPS',
    500: 'IKE',         514: 'Syslog',      515: 'LPD',
    546: 'DHCPv6-C',   547: 'DHCPv6-S',    548: 'AFP',
    587: 'SMTP-Sub',    631: 'IPP',         636: 'LDAPS',
    993: 'IMAPS',       995: 'POP3S',
    # VPN & tunnels
    1080: 'SOCKS',      1194: 'OpenVPN',    1701: 'L2TP',
    1723: 'PPTP',       4500: 'IKE-NAT',   51820: 'WireGuard',
    # Databases
    1433: 'MSSQL',      1521: 'Oracle',     3306: 'MySQL',
    5432: 'PostgreSQL', 6379: 'Redis',      7000: 'Cassandra',
    7001: 'Cassandra',  9200: 'Elastic',    9300: 'Elastic',
    11211: 'Memcached', 27017: 'MongoDB',   27018: 'MongoDB',
    # Message brokers & streaming
    1883: 'MQTT',       4369: 'EPMD',       5672: 'AMQP',
    8883: 'MQTT-TLS',   9092: 'Kafka',      15672: 'RabbitMQ',
    # Containers & orchestration
    2375: 'Docker',     2376: 'Docker-TLS', 2181: 'ZooKeeper',
    6443: 'k8s-API',    10250: 'kubelet',   10255: 'kubelet-ro',
    # Dev & monitoring
    3000: 'Grafana/Dev',4000: 'Dev',        5000: 'Flask/Dev',
    5601: 'Kibana',     8080: 'HTTP-Alt',   8081: 'HTTP-Alt',
    8443: 'HTTPS-Alt',  8888: 'Jupyter',    9000: 'Sonar/Dev',
    9090: 'Prometheus', 10000: 'Webmin',    50000: 'SAP',
    # Apple ecosystem
    5353: 'mDNS',       7070: 'RTSP',       49152: 'AirPlay',
    5000: 'AirPlay-2',
    # Remote desktop & screen sharing
    3389: 'RDP',        5900: 'VNC',        5901: 'VNC-1',
    # Media & streaming
    32400: 'Plex',      8096: 'Jellyfin',   8920: 'Jellyfin-TLS',
    # Misc
    2049: 'NFS',        3478: 'STUN',       3690: 'SVN',
    5222: 'XMPP',       5228: 'FCM/Android',8008: 'HTTP-Alt',
}

STYLESHEET = f"""
* {{
    font-family: 'Menlo', 'Monaco', 'Courier New', monospace;
}}
QMainWindow {{
    background-color: {C['bg']};
}}
QWidget {{
    background-color: transparent;
    color: {C['white']};
}}
QTabWidget::pane {{
    border: 2px solid {C['border']};
    background: {C['panel']};
    border-radius: 6px;
    margin-top: -1px;
}}
QTabBar::tab {{
    background: {C['card']};
    color: {C['dim']};
    padding: 7px 18px;
    border: 1px solid {C['border']};
    border-bottom: none;
    border-radius: 4px 4px 0 0;
    margin-right: 3px;
    font-weight: bold;
    font-size: 11px;
    min-width: 80px;
}}
QTabBar::tab:selected {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 {C['violet']}, stop:1 {C['panel']});
    color: {C['yellow']};
    border: 1px solid {C['purple']};
    border-bottom: 1px solid {C['panel']};
}}
QTabBar::tab:hover:!selected {{
    background: {C['border']};
    color: {C['white']};
}}
QTableWidget {{
    background: {C['panel']};
    color: {C['white']};
    gridline-color: {C['border']};
    border: 1px solid {C['border']};
    font-size: 11px;
    alternate-background-color: {C['card']};
    selection-background-color: {C['violet']};
    selection-color: {C['yellow']};
}}
QTableWidget::item {{
    padding: 3px 6px;
    border: none;
}}
QTableWidget::item:selected {{
    background: {C['violet']};
    color: {C['yellow']};
}}
QHeaderView::section {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #3B0080, stop:1 {C['card']});
    color: {C['cyan']};
    padding: 5px 8px;
    border: 1px solid {C['border']};
    font-weight: bold;
    font-size: 10px;
    text-transform: uppercase;
}}
QPushButton {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 {C['violet']}, stop:1 #0D001F);
    color: {C['white']};
    border: 1px solid {C['purple']};
    border-radius: 5px;
    padding: 5px 14px;
    font-weight: bold;
    font-size: 11px;
}}
QPushButton:hover {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 {C['purple']}, stop:1 {C['violet']});
    border: 1px solid {C['cyan']};
    color: {C['yellow']};
}}
QPushButton:pressed {{
    background: {C['card']};
    border: 1px solid {C['pink']};
    padding-top: 6px;
}}
QPushButton#allow_btn {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #007A3D, stop:1 #004020);
    border: 2px solid {C['allow']};
    color: {C['allow']};
    font-size: 15px;
    padding: 10px 28px;
    border-radius: 8px;
    font-weight: bold;
}}
QPushButton#allow_btn:hover {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 {C['allow']}, stop:1 #007A3D);
    color: #000;
}}
QPushButton#deny_btn {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 #880000, stop:1 #3A0000);
    border: 2px solid {C['deny']};
    color: {C['deny']};
    font-size: 15px;
    padding: 10px 28px;
    border-radius: 8px;
    font-weight: bold;
}}
QPushButton#deny_btn:hover {{
    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
        stop:0 {C['deny']}, stop:1 #880000);
    color: #fff;
}}
QPushButton#skip_btn {{
    background: transparent;
    border: 1px dashed {C['dim']};
    color: {C['dim']};
    font-size: 10px;
    padding: 4px 12px;
}}
QPushButton#skip_btn:hover {{
    border-color: {C['orange']};
    color: {C['orange']};
}}
QComboBox {{
    background: {C['card']};
    color: {C['white']};
    border: 1px solid {C['border']};
    border-radius: 4px;
    padding: 4px 8px;
    font-size: 11px;
    min-width: 160px;
}}
QComboBox:focus {{ border-color: {C['cyan']}; }}
QComboBox::drop-down {{ border: none; width: 22px; }}
QComboBox::down-arrow {{
    width: 10px; height: 10px;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 6px solid {C['purple']};
}}
QComboBox QAbstractItemView {{
    background: {C['card']};
    color: {C['white']};
    selection-background-color: {C['violet']};
    selection-color: {C['yellow']};
    border: 1px solid {C['border']};
    padding: 4px;
}}
QLineEdit {{
    background: {C['card']};
    color: {C['white']};
    border: 1px solid {C['border']};
    border-radius: 4px;
    padding: 4px 8px;
    font-size: 11px;
}}
QLineEdit:focus {{ border-color: {C['cyan']}; }}
QLineEdit::placeholder {{ color: {C['dim']}; }}
QTextEdit {{
    background: {C['panel']};
    color: {C['lime']};
    border: 1px solid {C['border']};
    font-size: 10px;
    border-radius: 4px;
}}
QScrollBar:vertical {{
    background: {C['card']};
    width: 8px;
    border-radius: 4px;
    margin: 0;
}}
QScrollBar::handle:vertical {{
    background: {C['violet']};
    border-radius: 4px;
    min-height: 20px;
}}
QScrollBar::handle:vertical:hover {{ background: {C['purple']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: {C['card']};
    height: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {C['violet']};
    border-radius: 4px;
}}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}
QGroupBox {{
    border: 1px solid {C['border']};
    border-radius: 8px;
    margin-top: 14px;
    padding-top: 6px;
    font-weight: bold;
    font-size: 11px;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 5px;
    color: {C['cyan']};
}}
QLabel {{ background: transparent; }}
QCheckBox {{
    color: {C['white']};
    spacing: 6px;
    font-size: 11px;
}}
QCheckBox::indicator {{
    width: 15px; height: 15px;
    border: 1px solid {C['border']};
    border-radius: 3px;
    background: {C['card']};
}}
QCheckBox::indicator:checked {{
    background: {C['violet']};
    border-color: {C['purple']};
}}
QMenu {{
    background: {C['card']};
    color: {C['white']};
    border: 1px solid {C['border']};
    padding: 4px;
}}
QMenu::item {{ padding: 4px 20px; border-radius: 3px; }}
QMenu::item:selected {{ background: {C['violet']}; color: {C['yellow']}; }}
QSplitter::handle {{
    background: {C['border']};
    height: 4px;
}}
QSplitter::handle:hover {{
    background: {C['purple']};
}}
QProgressBar {{
    background: {C['card']};
    border: 1px solid {C['border']};
    border-radius: 4px;
    height: 12px;
    text-align: center;
    font-size: 9px;
}}
QProgressBar::chunk {{
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 {C['violet']}, stop:1 {C['magenta']});
    border-radius: 3px;
}}
QStatusBar {{
    background: {C['card']};
    color: {C['lime']};
    border-top: 1px solid {C['border']};
    font-size: 10px;
}}
"""


# ══════════════════════════════════════════════════════════════════════
#  📊  DATA STRUCTURES
# ══════════════════════════════════════════════════════════════════════

@dataclass
class ConnInfo:
    pid: int
    process: str
    local_ip: str
    local_port: int
    remote_ip: str
    remote_port: int
    proto: str
    status: str
    direction: str
    timestamp: str

    def key(self):
        return f"{self.process}|{self.remote_ip}|{self.remote_port}"

    def service(self):
        return KNOWN_PORTS.get(self.remote_port, KNOWN_PORTS.get(self.local_port, '—'))

    def remote_str(self):
        return f"{self.remote_ip}:{self.remote_port}"


@dataclass
class Rule:
    id: str
    process: str       # '*' = any
    remote_ip: str     # '*' = any
    remote_port: int   # 0 = any
    action: str        # 'allow' | 'deny'
    duration: str      # 'forever'|'3h'|'24h'|'1w'|'logout'
    created_at: str
    expires_at: Optional[str]
    note: str
    use_count: int

    def matches(self, c: ConnInfo) -> bool:
        if self.process != '*' and self.process.lower() not in c.process.lower():
            return False
        if self.remote_ip != '*' and self.remote_ip != c.remote_ip:
            return False
        if self.remote_port != 0 and self.remote_port != c.remote_port:
            return False
        return True

    def is_expired(self) -> bool:
        # 'logout' rules are cleaned on startup, never expire mid-session
        if self.duration in ('forever', 'logout') or not self.expires_at:
            return False
        return datetime.now() > datetime.fromisoformat(self.expires_at)

    def time_left_str(self) -> str:
        """Human-readable countdown to expiry."""
        if self.duration == 'forever':
            return '∞ Forever'
        if self.duration == 'logout':
            return '⬛ Until logout'
        if not self.expires_at:
            return '—'
        delta = datetime.fromisoformat(self.expires_at) - datetime.now()
        total = int(delta.total_seconds())
        if total <= 0:
            return '⏰ Expired'
        hours, rem = divmod(total, 3600)
        mins = rem // 60
        if hours >= 24:
            days = hours // 24
            return f'{days}d {hours % 24}h left'
        if hours > 0:
            return f'{hours}h {mins}m left'
        return f'{mins}m left'


# ══════════════════════════════════════════════════════════════════════
#  🗄️  DATABASE
# ══════════════════════════════════════════════════════════════════════

class Database:
    def __init__(self):
        path = Path.home() / '.nerdathron3000' / 'traffic.db'
        path.parent.mkdir(parents=True, exist_ok=True)
        self.path = str(path)
        self._lock = threading.Lock()
        self._conn = sqlite3.connect(self.path, check_same_thread=False)
        self._init()

    def _init(self):
        c = self._conn.cursor()
        c.executescript('''
            CREATE TABLE IF NOT EXISTS log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                ts          TEXT,
                process     TEXT,
                pid         INTEGER,
                local_ip    TEXT,
                local_port  INTEGER,
                remote_ip   TEXT,
                remote_port INTEGER,
                proto       TEXT,
                direction   TEXT,
                action      TEXT,
                service     TEXT
            );
            CREATE TABLE IF NOT EXISTS rules (
                id          TEXT PRIMARY KEY,
                process     TEXT,
                remote_ip   TEXT,
                remote_port INTEGER,
                action      TEXT,
                duration    TEXT,
                created_at  TEXT,
                expires_at  TEXT,
                note        TEXT,
                use_count   INTEGER DEFAULT 0
            );
            CREATE INDEX IF NOT EXISTS idx_log_ts ON log(ts);
            CREATE INDEX IF NOT EXISTS idx_log_proc ON log(process);
        ''')
        self._conn.commit()

    def log(self, c: ConnInfo, action: str):
        with self._lock:
            self._conn.execute(
                'INSERT INTO log VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?)',
                (c.timestamp, c.process, c.pid, c.local_ip, c.local_port,
                 c.remote_ip, c.remote_port, c.proto, c.direction, action, c.service()))
            self._conn.commit()

    def save_rule(self, r: Rule):
        with self._lock:
            self._conn.execute(
                'INSERT OR REPLACE INTO rules VALUES (?,?,?,?,?,?,?,?,?,?)',
                (r.id, r.process, r.remote_ip, r.remote_port, r.action,
                 r.duration, r.created_at, r.expires_at, r.note, r.use_count))
            self._conn.commit()

    def delete_rule(self, rule_id: str):
        with self._lock:
            self._conn.execute('DELETE FROM rules WHERE id=?', (rule_id,))
            self._conn.commit()

    def delete_logout_rules(self):
        """Called once on startup to clear rules that only last until logout."""
        with self._lock:
            self._conn.execute("DELETE FROM rules WHERE duration='logout'")
            self._conn.commit()

    def load_rules(self) -> List[Rule]:
        with self._lock:
            rows = self._conn.execute('SELECT * FROM rules').fetchall()
        result = []
        for r in rows:
            result.append(Rule(
                id=r[0], process=r[1], remote_ip=r[2], remote_port=r[3],
                action=r[4], duration=r[5], created_at=r[6], expires_at=r[7],
                note=r[8], use_count=r[9]
            ))
        return result

    def get_log(self, limit=500, filter_text='') -> list:
        with self._lock:
            if filter_text:
                q = f'%{filter_text}%'
                return self._conn.execute(
                    '''SELECT ts, process, pid, local_ip, local_port, remote_ip,
                              remote_port, proto, direction, action, service
                       FROM log WHERE process LIKE ? OR remote_ip LIKE ?
                       ORDER BY id DESC LIMIT ?''', (q, q, limit)).fetchall()
            return self._conn.execute(
                '''SELECT ts, process, pid, local_ip, local_port, remote_ip,
                          remote_port, proto, direction, action, service
                   FROM log ORDER BY id DESC LIMIT ?''', (limit,)).fetchall()

    def get_all_log(self) -> list:
        """Fetch full log for xlsx export."""
        with self._lock:
            return self._conn.execute(
                '''SELECT ts, process, pid, local_ip, local_port, remote_ip,
                          remote_port, proto, direction, action, service
                   FROM log ORDER BY id DESC''').fetchall()

    def stats(self) -> dict:
        with self._lock:
            ex = self._conn.execute
            return {
                'total':         ex('SELECT COUNT(*) FROM log').fetchone()[0],
                'allowed':       ex("SELECT COUNT(*) FROM log WHERE action='allow'").fetchone()[0],
                'denied':        ex("SELECT COUNT(*) FROM log WHERE action='deny'").fetchone()[0],
                'top_procs':     ex('SELECT process, COUNT(*) c FROM log GROUP BY process ORDER BY c DESC LIMIT 8').fetchall(),
                'top_ips':       ex("SELECT remote_ip, COUNT(*) c FROM log WHERE remote_ip!='' GROUP BY remote_ip ORDER BY c DESC LIMIT 8").fetchall(),
                'top_ports':     ex('SELECT remote_port, COUNT(*) c FROM log GROUP BY remote_port ORDER BY c DESC LIMIT 8').fetchall(),
                'recent_denies': ex("SELECT ts, process, remote_ip, remote_port FROM log WHERE action='deny' ORDER BY id DESC LIMIT 5").fetchall(),
            }


# ══════════════════════════════════════════════════════════════════════
#  ⚙️  RULE ENGINE
# ══════════════════════════════════════════════════════════════════════

class RuleEngine:
    def __init__(self, db: Database):
        self.db = db
        self._lock = threading.Lock()
        self.rules: List[Rule] = []
        # Clean logout rules from previous session before loading
        self.db.delete_logout_rules()
        self.reload()

    def reload(self):
        with self._lock:
            self.rules = [r for r in self.db.load_rules() if not r.is_expired()]

    def check(self, c: ConnInfo) -> Optional[str]:
        with self._lock:
            for r in self.rules:
                if r.is_expired():
                    continue
                if r.matches(c):
                    r.use_count += 1
                    return r.action
        return None

    def add(self, r: Rule):
        with self._lock:
            self.rules.append(r)
        self.db.save_rule(r)

    def remove(self, rule_id: str):
        with self._lock:
            self.rules = [r for r in self.rules if r.id != rule_id]
        self.db.delete_rule(rule_id)

    def get_all(self) -> List[Rule]:
        with self._lock:
            return list(self.rules)


# ══════════════════════════════════════════════════════════════════════
#  🔄  CONNECTION MONITOR THREAD
# ══════════════════════════════════════════════════════════════════════

class Monitor(QThread):
    new_conn  = pyqtSignal(object)
    all_conns = pyqtSignal(list)

    def __init__(self, engine: RuleEngine):
        super().__init__()
        self.engine = engine
        self.running = True
        self.paused = False
        self.seen = set()

    def run(self):
        while self.running:
            if not self.paused:
                try:
                    snapshot = []
                    for c in psutil.net_connections(kind='inet'):
                        if not c.raddr:
                            continue
                        try:
                            proc = psutil.Process(c.pid).name() if c.pid else 'kernel'
                        except (psutil.NoSuchProcess, psutil.AccessDenied):
                            proc = f'pid:{c.pid or "?"}'

                        li = c.laddr.ip if c.laddr else ''
                        lp = c.laddr.port if c.laddr else 0
                        ri = c.raddr.ip
                        rp = c.raddr.port

                        if ri.startswith('127.') or ri == '::1':
                            continue

                        direction = 'outbound' if lp > 1024 else 'inbound'
                        proto = 'TCP' if c.type == socket.SOCK_STREAM else 'UDP'

                        info = ConnInfo(
                            pid=c.pid or 0,
                            process=proc,
                            local_ip=li, local_port=lp,
                            remote_ip=ri, remote_port=rp,
                            proto=proto,
                            status=c.status or '—',
                            direction=direction,
                            timestamp=datetime.now().isoformat(timespec='seconds')
                        )
                        snapshot.append(info)

                        k = info.key()
                        if k not in self.seen:
                            self.seen.add(k)
                            self.new_conn.emit(info)

                    self.all_conns.emit(snapshot)
                except Exception:
                    pass
            self.msleep(1200)

    def stop(self):
        self.running = False
        self.quit()


# ══════════════════════════════════════════════════════════════════════
#  🚨  ALERT DIALOG  — redesigned layout so scope never hovers Deny
#     • Allow is LEFT and is the default (Enter key)
#     • Deny is RIGHT
#     • Scope + Duration sit in their own row, clearly separated
#     • 4th scope option: This Port, Any IP
#     • Duration: 3h / 24h / 1 week / Forever / Until logout
# ══════════════════════════════════════════════════════════════════════

class AlertDialog(QDialog):
    def __init__(self, conn: ConnInfo, parent=None):
        super().__init__(parent)
        self.conn = conn
        self.result_action = None
        self.result_rule = None

        self.setWindowTitle("⚠ NEW CONNECTION ALERT")
        self.setFixedSize(580, 500)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.CustomizeWindowHint |
            Qt.WindowType.WindowTitleHint
        )
        self.setStyleSheet(STYLESHEET)
        self._build()
        self._flash_on = True
        self._flash_timer = QTimer(self)
        self._flash_timer.timeout.connect(self._flash)
        self._flash_timer.start(550)

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(18, 18, 18, 18)

        # — Banner —
        banner = QLabel("🚨  UNKNOWN CONNECTION INTERCEPTED  🚨")
        banner.setAlignment(Qt.AlignmentFlag.AlignCenter)
        banner.setStyleSheet(f"""
            font-size: 13px; font-weight: bold; color: {C['pink']};
            background: {C['card']};
            border: 1px solid {C['pink']};
            border-radius: 6px; padding: 9px;
        """)
        layout.addWidget(banner)

        # — Process name —
        proc_lbl = QLabel(f"🖥  {self.conn.process}")
        proc_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        proc_lbl.setStyleSheet(f"""
            font-size: 22px; font-weight: bold; color: {C['yellow']};
            padding: 4px;
        """)
        layout.addWidget(proc_lbl)

        # — Info grid —
        info = QFrame()
        info.setStyleSheet(f"""
            QFrame {{
                background: {C['card']};
                border: 1px solid {C['border']};
                border-radius: 8px;
            }}
        """)
        grid = QVBoxLayout(info)
        grid.setSpacing(4)
        grid.setContentsMargins(14, 8, 14, 8)

        def row(icon, label, value, color):
            h = QHBoxLayout()
            ico = QLabel(icon); ico.setFixedWidth(22)
            lbl = QLabel(label)
            lbl.setStyleSheet(f"color:{C['dim']};font-size:10px;font-weight:bold;min-width:100px;")
            val = QLabel(str(value))
            val.setStyleSheet(f"color:{color};font-size:12px;font-weight:bold;")
            val.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
            val.setCursor(QCursor(Qt.CursorShape.IBeamCursor))
            h.addWidget(ico); h.addWidget(lbl); h.addWidget(val); h.addStretch()
            grid.addLayout(h)

        svc = self.conn.service()
        port_str = f"{self.conn.remote_port}  ·  [{svc}]" if svc != '—' else str(self.conn.remote_port)
        dir_icon = "⬆" if self.conn.direction == 'outbound' else "⬇"
        dir_color = C['orange'] if self.conn.direction == 'outbound' else C['pink']

        row("🌐", "Remote IP",   self.conn.remote_ip,  C['cyan'])
        row("🔌", "Port / Svc",  port_str,              C['lime'])
        row("📡", "Direction",   f"{dir_icon} {self.conn.direction.upper()}", dir_color)
        row("📦", "Protocol",    self.conn.proto,       C['purple'])
        row("🔢", "PID",         self.conn.pid,         C['white'])
        row("🕐", "Time",        self.conn.timestamp,   C['dim'])

        layout.addWidget(info)

        # ── Rule Options: scope LEFT, duration RIGHT, clearly separated ──
        opts = QGroupBox("  Rule Settings")
        opts_layout = QHBoxLayout(opts)
        opts_layout.setSpacing(20)

        # Scope
        scope_v = QVBoxLayout()
        scope_lbl = QLabel("🔭  Scope:")
        scope_lbl.setStyleSheet(f"color:{C['dim']};font-size:10px;font-weight:bold;")
        self.scope = QComboBox()
        self.scope.addItems([
            f"This IP  ({self.conn.remote_ip})",
            f"This IP + Port  (:{self.conn.remote_port})",
            "Any IP  (wildcard)",
            f"This Port, Any IP  (:{self.conn.remote_port})",
        ])
        scope_v.addWidget(scope_lbl)
        scope_v.addWidget(self.scope)

        # Separator
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet(f"color:{C['border']};")

        # Duration
        dur_v = QVBoxLayout()
        dur_lbl = QLabel("⏱  Duration:")
        dur_lbl.setStyleSheet(f"color:{C['dim']};font-size:10px;font-weight:bold;")
        self.dur = QComboBox()
        self.dur.addItems(["3 Hours", "24 Hours", "1 Week", "Forever ∞", "Until logout"])
        self.dur.setCurrentIndex(3)  # default: Forever
        dur_v.addWidget(dur_lbl)
        dur_v.addWidget(self.dur)

        opts_layout.addLayout(scope_v)
        opts_layout.addWidget(sep)
        opts_layout.addLayout(dur_v)
        layout.addWidget(opts)

        # ── Buttons: ALLOW left (default/Enter), DENY right ──
        # This means scope dropdown is never hovering above Deny
        btn_row = QHBoxLayout()
        btn_row.setSpacing(16)

        self.allow_btn = QPushButton("✅  ALLOW")
        self.allow_btn.setObjectName("allow_btn")
        self.allow_btn.setMinimumHeight(48)
        self.allow_btn.setDefault(True)       # ← Enter key fires this
        self.allow_btn.setAutoDefault(True)
        self.allow_btn.clicked.connect(self._allow)

        self.deny_btn = QPushButton("🚫  DENY")
        self.deny_btn.setObjectName("deny_btn")
        self.deny_btn.setMinimumHeight(48)
        self.deny_btn.setDefault(False)
        self.deny_btn.setAutoDefault(False)
        self.deny_btn.clicked.connect(self._deny)

        btn_row.addWidget(self.allow_btn)
        btn_row.addWidget(self.deny_btn)
        layout.addLayout(btn_row)

        skip_btn = QPushButton("⏭  Allow once (no rule created)")
        skip_btn.setObjectName("skip_btn")
        skip_btn.clicked.connect(self._skip)
        layout.addWidget(skip_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        # Set initial focus to allow button so Enter works immediately
        self.allow_btn.setFocus()

    def _make_rule(self, action: str) -> Rule:
        scope_i = self.scope.currentIndex()
        dur_i   = self.dur.currentIndex()

        # Duration mapping: 3h / 24h / 1w / forever / logout
        dur_keys = ['3h', '24h', '1w', 'forever', 'logout']
        dur_mins = [180,  1440,  10080, 0,        0]
        duration = dur_keys[dur_i]

        # Scope mapping
        # 0: this IP, any port
        # 1: this IP + this port
        # 2: any IP, any port
        # 3: any IP, this port
        if scope_i == 0:
            remote_ip, remote_port = self.conn.remote_ip, 0
        elif scope_i == 1:
            remote_ip, remote_port = self.conn.remote_ip, self.conn.remote_port
        elif scope_i == 2:
            remote_ip, remote_port = '*', 0
        else:  # scope_i == 3
            remote_ip, remote_port = '*', self.conn.remote_port

        expires_at = None
        if duration not in ('forever', 'logout') and dur_mins[dur_i]:
            expires_at = (datetime.now() + timedelta(minutes=dur_mins[dur_i])).isoformat()

        return Rule(
            id=uuid.uuid4().hex[:8],
            process=self.conn.process,
            remote_ip=remote_ip,
            remote_port=remote_port,
            action=action,
            duration=duration,
            created_at=datetime.now().isoformat(timespec='seconds'),
            expires_at=expires_at,
            note='',
            use_count=0
        )

    def _allow(self):
        self.result_action = 'allow'
        self.result_rule   = self._make_rule('allow')
        self.accept()

    def _deny(self):
        self.result_action = 'deny'
        self.result_rule   = self._make_rule('deny')
        self.accept()

    def _skip(self):
        self.result_action = 'allow'
        self.result_rule   = None
        self.accept()

    def _flash(self):
        self._flash_on = not self._flash_on
        c = C['pink'] if self._flash_on else C['deny']
        self.setStyleSheet(STYLESHEET + f"""
            QDialog {{
                background: {C['bg']};
                border: 3px solid {c};
            }}
        """)

    def keyPressEvent(self, event):
        """Ensure Enter always triggers Allow, Escape does nothing dangerous."""
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            self._allow()
        elif event.key() == Qt.Key.Key_Escape:
            self._skip()
        else:
            super().keyPressEvent(event)


# ══════════════════════════════════════════════════════════════════════
#  ⚡  LIVE TRAFFIC PANEL  (columns now reorderable)
# ══════════════════════════════════════════════════════════════════════

class LivePanel(QWidget):
    def __init__(self):
        super().__init__()
        lay = QVBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(4)

        header = QHBoxLayout()
        title = QLabel("⚡  LIVE CONNECTIONS")
        title.setStyleSheet(f"color:{C['lime']};font-size:12px;font-weight:bold;")
        self.count_lbl = QLabel("— active")
        self.count_lbl.setStyleSheet(f"color:{C['cyan']};font-size:10px;")
        self.filter_box = QLineEdit()
        self.filter_box.setPlaceholderText("Filter…")
        self.filter_box.setFixedWidth(140)
        self.filter_box.textChanged.connect(self._filter)
        header.addWidget(title)
        header.addStretch()
        header.addWidget(self.count_lbl)
        header.addWidget(self.filter_box)
        lay.addLayout(header)

        self.table = QTableWidget()
        cols = ["Process", "PID", "Direction", "Remote IP", "Port", "Service", "Proto", "Status"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(True)          # ← drag to reorder columns
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setShowGrid(True)
        lay.addWidget(self.table)

        self._all: List[ConnInfo] = []

    def update(self, conns: list):
        self._all = conns
        self.count_lbl.setText(f"{len(conns)} active")
        self._filter(self.filter_box.text())

    def _filter(self, text: str):
        data = self._all
        if text:
            data = [c for c in data if text.lower() in c.process.lower()
                    or text.lower() in c.remote_ip.lower()]
        self.table.setRowCount(0)
        for c in data:
            r = self.table.rowCount()
            self.table.insertRow(r)
            di = "⬆ OUT" if c.direction == 'outbound' else "⬇ IN"
            dc = C['orange'] if c.direction == 'outbound' else C['pink']
            for col, (txt, clr) in enumerate([
                (c.process, C['yellow']), (str(c.pid), C['dim']),
                (di, dc), (c.remote_ip, C['cyan']),
                (str(c.remote_port), C['lime']), (c.service(), C['gold']),
                (c.proto, C['purple']), (c.status, C['dim'])
            ]):
                item = QTableWidgetItem(txt)
                item.setForeground(QColor(clr))
                self.table.setItem(r, col, item)


# ══════════════════════════════════════════════════════════════════════
#  📋  RULES PANEL  (with live countdown + reorderable columns)
# ══════════════════════════════════════════════════════════════════════

class RulesPanel(QWidget):
    def __init__(self, engine: RuleEngine):
        super().__init__()
        self.engine = engine
        lay = QVBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(4)

        header = QHBoxLayout()
        title = QLabel("📋  ACTIVE RULES")
        title.setStyleSheet(f"color:{C['orange']};font-size:12px;font-weight:bold;")
        ref_btn = QPushButton("🔄 Refresh")
        ref_btn.clicked.connect(self.refresh)
        del_btn = QPushButton("🗑 Delete Selected")
        del_btn.setStyleSheet(f"background:#3A0000;border-color:{C['deny']};color:{C['deny']};")
        del_btn.clicked.connect(self._delete)
        header.addWidget(title)
        header.addStretch()
        header.addWidget(ref_btn)
        header.addWidget(del_btn)
        lay.addLayout(header)

        self.table = QTableWidget()
        cols = ["ID", "Process", "Remote IP", "Port", "Action", "Duration", "Time Left", "Created", "Uses", "Note"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(True)          # ← drag to reorder columns
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lay.addWidget(self.table)

        self._ids = []
        self.refresh()

        # Refresh time-left column every 30 seconds
        self._tick = QTimer(self)
        self._tick.timeout.connect(self._update_time_left)
        self._tick.start(30_000)

    def refresh(self):
        self.engine.reload()
        rules = self.engine.get_all()
        self.table.setRowCount(0)
        self._ids = []
        for rule in rules:
            r = self.table.rowCount()
            self.table.insertRow(r)
            self._ids.append(rule.id)
            ac = C['allow'] if rule.action == 'allow' else C['deny']
            ai = "✅" if rule.action == 'allow' else "🚫"
            tl = rule.time_left_str()
            tl_color = C['warn'] if 'left' in tl and 'h' not in tl else C['teal']
            if tl == '∞ Forever':
                tl_color = C['dim']
            elif 'logout' in tl:
                tl_color = C['blue']
            elif 'Expired' in tl:
                tl_color = C['deny']
            for col, (txt, clr) in enumerate([
                (rule.id,                                        C['dim']),
                (rule.process,                                   C['yellow']),
                (rule.remote_ip,                                 C['cyan']),
                (str(rule.remote_port) if rule.remote_port else '*', C['lime']),
                (f"{ai} {rule.action.upper()}",                  ac),
                (rule.duration,                                  C['orange']),
                (tl,                                             tl_color),
                (rule.created_at[:16],                           C['dim']),
                (str(rule.use_count),                            C['white']),
                (rule.note,                                      C['dim']),
            ]):
                item = QTableWidgetItem(txt)
                item.setForeground(QColor(clr))
                self.table.setItem(r, col, item)

    def _update_time_left(self):
        """Lightweight: only refresh the Time Left column (col 6)."""
        rules = self.engine.get_all()
        for row_idx, rule in enumerate(rules):
            if row_idx >= self.table.rowCount():
                break
            tl = rule.time_left_str()
            tl_color = C['warn'] if 'left' in tl and 'h' not in tl else C['teal']
            if tl == '∞ Forever':
                tl_color = C['dim']
            elif 'logout' in tl:
                tl_color = C['blue']
            elif 'Expired' in tl:
                tl_color = C['deny']
            item = QTableWidgetItem(tl)
            item.setForeground(QColor(tl_color))
            self.table.setItem(row_idx, 6, item)

    def _delete(self):
        rows = set(i.row() for i in self.table.selectedItems())
        for row in sorted(rows, reverse=True):
            if row < len(self._ids):
                self.engine.remove(self._ids[row])
        self.refresh()


# ══════════════════════════════════════════════════════════════════════
#  📓  LOG PANEL  (reorderable columns + xlsx export)
# ══════════════════════════════════════════════════════════════════════

class LogPanel(QWidget):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        lay = QVBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 6)
        lay.setSpacing(4)

        header = QHBoxLayout()
        title = QLabel("📓  TRAFFIC LOG")
        title.setStyleSheet(f"color:{C['cyan']};font-size:12px;font-weight:bold;")
        self.filter_box = QLineEdit()
        self.filter_box.setPlaceholderText("Filter by process or IP…")
        self.filter_box.setFixedWidth(180)
        self.filter_box.textChanged.connect(self._filter_changed)
        ref_btn = QPushButton("🔄")
        ref_btn.setFixedWidth(36)
        ref_btn.clicked.connect(self.refresh)
        lbl_count = QLabel("Last 500 entries")
        lbl_count.setStyleSheet(f"color:{C['dim']};font-size:10px;")

        xlsx_btn = QPushButton("📊 Export .xlsx")
        xlsx_btn.setToolTip("Export full log to Excel spreadsheet")
        xlsx_btn.clicked.connect(self._export_xlsx)
        if not XLSX_AVAILABLE:
            xlsx_btn.setEnabled(False)
            xlsx_btn.setToolTip("Install openpyxl to enable: pip install openpyxl")

        header.addWidget(title)
        header.addStretch()
        header.addWidget(lbl_count)
        header.addWidget(self.filter_box)
        header.addWidget(ref_btn)
        header.addWidget(xlsx_btn)
        lay.addLayout(header)

        self.table = QTableWidget()
        cols = ["Time", "Process", "PID", "Local", "Remote IP", "Port", "Proto", "Dir", "Action", "Svc"]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(True)          # ← drag to reorder columns
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lay.addWidget(self.table)

        self._filter_text = ''
        self.refresh()

    def refresh(self):
        self._populate(self.db.get_log(500, self._filter_text))

    def _filter_changed(self, text):
        self._filter_text = text
        self.refresh()

    def _populate(self, rows):
        self.table.setRowCount(0)
        for row in rows:
            ts, proc, pid, lip, lport, rip, rport, proto, direction, action, svc = row
            r = self.table.rowCount()
            self.table.insertRow(r)
            ac = C['allow'] if action == 'allow' else C['deny'] if action == 'deny' else C['warn']
            ai = "✅" if action == 'allow' else "🚫" if action == 'deny' else "⏳"
            di = "⬆" if direction == 'outbound' else "⬇"
            for col, (txt, clr) in enumerate([
                (ts[:19],          C['dim']),
                (proc,             C['yellow']),
                (str(pid),         C['dim']),
                (f"{lip}:{lport}", C['dim']),
                (rip,              C['cyan']),
                (str(rport),       C['lime']),
                (proto,            C['purple']),
                (di,               C['orange']),
                (f"{ai}{action}",  ac),
                (svc,              C['gold']),
            ]):
                item = QTableWidgetItem(txt)
                item.setForeground(QColor(clr))
                self.table.setItem(r, col, item)

    def add_live(self, c: ConnInfo, action: str):
        ac = C['allow'] if action == 'allow' else C['deny'] if action == 'deny' else C['warn']
        ai = "✅" if action == 'allow' else "🚫" if action == 'deny' else "⏳"
        di = "⬆" if c.direction == 'outbound' else "⬇"

        self.table.insertRow(0)
        for col, (txt, clr) in enumerate([
            (c.timestamp[:19],               C['dim']),
            (c.process,                      C['yellow']),
            (str(c.pid),                     C['dim']),
            (f"{c.local_ip}:{c.local_port}", C['dim']),
            (c.remote_ip,                    C['cyan']),
            (str(c.remote_port),             C['lime']),
            (c.proto,                        C['purple']),
            (di,                             C['orange']),
            (f"{ai}{action}",                ac),
            (c.service(),                    C['gold']),
        ]):
            item = QTableWidgetItem(txt)
            item.setForeground(QColor(clr))
            self.table.setItem(0, col, item)

        while self.table.rowCount() > 500:
            self.table.removeRow(self.table.rowCount() - 1)

    def _export_xlsx(self):
        """Export full traffic log to .xlsx file."""
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Log to Excel",
            str(Path.home() / f"nerdathron_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        try:
            rows = self.db.get_all_log()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Traffic Log"

            # Header row styling
            headers = ["Timestamp", "Process", "PID", "Local IP", "Local Port",
                       "Remote IP", "Remote Port", "Protocol", "Direction", "Action", "Service"]
            header_fill   = PatternFill("solid", fgColor="1A0040")
            header_font   = Font(bold=True, color="00F5FF", name="Courier New")
            header_align  = Alignment(horizontal="center")

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = header_align

            # Data rows
            allow_fill  = PatternFill("solid", fgColor="002210")
            deny_fill   = PatternFill("solid", fgColor="220000")
            allow_font  = Font(color="00FF88", name="Courier New", size=10)
            deny_font   = Font(color="FF3333", name="Courier New", size=10)
            normal_font = Font(color="C8B0FF", name="Courier New", size=10)

            for row_idx, row in enumerate(rows, 2):
                ts, proc, pid, lip, lport, rip, rport, proto, direction, action, svc = row
                values = [ts, proc, pid, lip, lport, rip, rport, proto, direction, action, svc or '—']
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if action == 'allow':
                        cell.fill = allow_fill
                        cell.font = allow_font
                    elif action == 'deny':
                        cell.fill = deny_fill
                        cell.font = deny_font
                    else:
                        cell.font = normal_font

            # Auto-fit columns
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            # Freeze top row
            ws.freeze_panes = "A2"

            wb.save(path)

            QMessageBox.information(
                self, "Export Complete",
                f"✅  Log exported successfully!\n\n{len(rows)} entries → {path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"❌  Export error:\n{e}")


# ══════════════════════════════════════════════════════════════════════
#  📊  STATS PANEL
# ══════════════════════════════════════════════════════════════════════

class StatsPanel(QWidget):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        lay = QVBoxLayout(self)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(8)

        title = QLabel("📊  NETWORK INTELLIGENCE REPORT")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet(f"""
            font-size: 14px; font-weight: bold; color: {C['yellow']};
            background: {C['card']}; border-radius: 6px; padding: 8px;
            border: 1px solid {C['border']};
        """)
        lay.addWidget(title)

        counters = QHBoxLayout()
        self._total   = self._counter("TOTAL",   "0", C['cyan'])
        self._allowed = self._counter("ALLOWED", "0", C['allow'])
        self._denied  = self._counter("DENIED",  "0", C['deny'])
        self._rules   = self._counter("RULES",   "0", C['orange'])
        for w in [self._total[0], self._allowed[0], self._denied[0], self._rules[0]]:
            counters.addWidget(w)
        lay.addLayout(counters)

        tables_row = QHBoxLayout()
        self._proc_tbl = self._mini_table("🖥  TOP PROCESSES",  ["Process", "Hits"])
        self._ip_tbl   = self._mini_table("🌐  TOP REMOTE IPs", ["IP Address", "Hits"])
        self._port_tbl = self._mini_table("🔌  TOP PORTS",      ["Port", "Hits"])
        tables_row.addWidget(self._proc_tbl[0])
        tables_row.addWidget(self._ip_tbl[0])
        tables_row.addWidget(self._port_tbl[0])
        lay.addLayout(tables_row)

        deny_group = QGroupBox("  🚫  RECENT DENIES")
        deny_layout = QVBoxLayout(deny_group)
        self._deny_txt = QTextEdit()
        self._deny_txt.setReadOnly(True)
        self._deny_txt.setMaximumHeight(100)
        deny_layout.addWidget(self._deny_txt)
        lay.addWidget(deny_group)

        refresh_btn = QPushButton("🔄  Refresh All Stats")
        refresh_btn.clicked.connect(self.refresh)
        lay.addWidget(refresh_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        self.refresh()

    def _counter(self, label, value, color):
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background: {C['card']};
                border: 2px solid {color};
                border-radius: 8px;
            }}
        """)
        fl = QVBoxLayout(frame)
        lbl = QLabel(label)
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet(f"color:{C['dim']};font-size:9px;font-weight:bold;letter-spacing:1px;")
        val = QLabel(value)
        val.setAlignment(Qt.AlignmentFlag.AlignCenter)
        val.setStyleSheet(f"color:{color};font-size:26px;font-weight:bold;")
        fl.addWidget(lbl); fl.addWidget(val)
        return frame, val

    def _mini_table(self, title, headers):
        frame = QFrame()
        frame.setStyleSheet(f"QFrame{{background:{C['panel']};border:1px solid {C['border']};border-radius:6px;}}")
        fl = QVBoxLayout(frame)
        fl.setContentsMargins(6, 6, 6, 6)
        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(f"color:{C['cyan']};font-size:11px;font-weight:bold;")
        fl.addWidget(title_lbl)
        tbl = QTableWidget()
        tbl.setColumnCount(2)
        tbl.setHorizontalHeaderLabels(headers)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        tbl.horizontalHeader().setSectionsMovable(True)
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setMaximumHeight(200)
        tbl.setAlternatingRowColors(True)
        fl.addWidget(tbl)
        return frame, tbl

    def refresh(self):
        s = self.db.stats()
        self._total[1].setText(str(s['total']))
        self._allowed[1].setText(str(s['allowed']))
        self._denied[1].setText(str(s['denied']))

        def fill(widget, data, colors):
            widget.setRowCount(0)
            for row in data:
                r = widget.rowCount()
                widget.insertRow(r)
                for col, (val, clr) in enumerate(zip(row, colors)):
                    item = QTableWidgetItem(str(val))
                    item.setForeground(QColor(clr))
                    widget.setItem(r, col, item)

        fill(self._proc_tbl[1], s['top_procs'],  [C['yellow'], C['cyan']])
        fill(self._ip_tbl[1],   s['top_ips'],    [C['orange'], C['cyan']])
        fill(self._port_tbl[1], s['top_ports'],  [C['lime'],   C['cyan']])

        deny_lines = []
        for ts, proc, ip, port in s['recent_denies']:
            deny_lines.append(f"[{ts[:16]}]  {proc}  →  {ip}:{port}")
        self._deny_txt.setPlainText('\n'.join(deny_lines) if deny_lines else '  (none yet)')


# ══════════════════════════════════════════════════════════════════════
#  🏠  MAIN WINDOW  — with adjustable QSplitter instead of fixed heights
# ══════════════════════════════════════════════════════════════════════

class Nerdathron3001(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🌈  NETWORK NERDATHRON 3001™  🌈")
        self.setMinimumSize(1200, 800)
        self.resize(1350, 900)
        self.setStyleSheet(STYLESHEET)

        self.db = Database()
        self.engine = RuleEngine(self.db)
        self._pending = False
        self._alert_count = 0
        self._allow_count = 0
        self._deny_count  = 0

        self._build_ui()
        self._setup_tray()
        self._start_monitor()
        self._setup_timers()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        ml = QVBoxLayout(central)
        ml.setSpacing(0)
        ml.setContentsMargins(0, 0, 0, 0)

        ml.addWidget(self._make_header())

        content = QWidget()
        content.setStyleSheet(f"background:{C['bg']};")
        cl = QVBoxLayout(content)
        cl.setContentsMargins(10, 8, 10, 10)
        cl.setSpacing(0)

        # ── Adjustable splitter between live connections and tabs ──
        self.splitter = QSplitter(Qt.Orientation.Vertical)
        self.splitter.setHandleWidth(6)
        self.splitter.setChildrenCollapsible(False)

        self.live_panel = LivePanel()
        self.tabs = QTabWidget()

        self.rules_panel = RulesPanel(self.engine)
        self.log_panel   = LogPanel(self.db)
        self.stats_panel = StatsPanel(self.db)

        self.tabs.addTab(self.rules_panel, "📋  RULES")
        self.tabs.addTab(self.log_panel,   "📓  LOG")
        self.tabs.addTab(self.stats_panel, "📊  STATS")

        self.splitter.addWidget(self.live_panel)
        self.splitter.addWidget(self.tabs)
        self.splitter.setSizes([240, 580])   # Initial proportions — user can drag

        cl.addWidget(self.splitter)
        ml.addWidget(content)

        sb = self.statusBar()
        sb.showMessage("🟢  NERDATHRON 3001 is online and watching the cosmos…  Stay groovy!")

    def _make_header(self):
        container = QWidget()
        container.setFixedHeight(72)
        container.setStyleSheet(f"""
            QWidget {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #2A0055, stop:0.25 #7200CA,
                    stop:0.5 #3D0070, stop:0.75 #7200CA, stop:1 #2A0055);
                border-bottom: 2px solid {C['pink']};
            }}
        """)
        lay = QHBoxLayout(container)
        lay.setContentsMargins(14, 8, 14, 8)

        logo = QLabel("🌈  NETWORK NERDATHRON 3001™")
        logo.setStyleSheet(f"""
            font-size: 21px; font-weight: bold; color: {C['yellow']};
            font-family: 'Courier New', monospace;
            background: transparent;
        """)

        self.monitor_lbl = QLabel("● MONITORING")
        self.monitor_lbl.setStyleSheet(f"color:{C['lime']};font-size:11px;font-weight:bold;background:transparent;")

        self.alert_lbl = QLabel("Alerts: 0")
        self.alert_lbl.setStyleSheet(f"color:{C['orange']};font-size:10px;background:transparent;")

        self.allow_lbl = QLabel("✅ 0")
        self.allow_lbl.setStyleSheet(f"color:{C['allow']};font-size:10px;background:transparent;")

        self.deny_lbl_hdr = QLabel("🚫 0")
        self.deny_lbl_hdr.setStyleSheet(f"color:{C['deny']};font-size:10px;background:transparent;")

        self.pause_btn = QPushButton("⏸ PAUSE")
        self.pause_btn.setFixedWidth(90)
        self.pause_btn.setStyleSheet("background:transparent;border:1px solid #BF5FFF;color:#F5EEFF;font-size:10px;padding:4px;")
        self.pause_btn.clicked.connect(self._toggle_monitor)

        clear_btn = QPushButton("🔁 Reset Seen")
        clear_btn.setFixedWidth(100)
        clear_btn.setStyleSheet("background:transparent;border:1px solid #BF5FFF;color:#F5EEFF;font-size:10px;padding:4px;")
        clear_btn.clicked.connect(self._reset_seen)
        clear_btn.setToolTip("Clears the list of known connections — all traffic will be re-evaluated")

        lay.addWidget(logo)
        lay.addStretch()
        lay.addWidget(self.monitor_lbl)
        lay.addSpacing(14)
        lay.addWidget(self.alert_lbl)
        lay.addSpacing(8)
        lay.addWidget(self.allow_lbl)
        lay.addSpacing(4)
        lay.addWidget(self.deny_lbl_hdr)
        lay.addSpacing(14)
        lay.addWidget(self.pause_btn)
        lay.addWidget(clear_btn)

        return container

    def _setup_tray(self):
        px = QPixmap(32, 32)
        px.fill(QColor(C['violet']))
        p = QPainter(px)
        p.setPen(QColor(C['lime']))
        f = QFont('Courier New', 8, QFont.Weight.Bold)
        p.setFont(f)
        p.drawText(1, 12, "N3K")
        p.end()

        self.tray = QSystemTrayIcon(QIcon(px), self)
        menu = QMenu()
        menu.addAction("Show Nerdathron").triggered.connect(self.show)
        menu.addSeparator()
        menu.addAction("Quit").triggered.connect(QApplication.quit)
        self.tray.setContextMenu(menu)
        self.tray.activated.connect(
            lambda reason: self.show() if reason == QSystemTrayIcon.ActivationReason.DoubleClick else None
        )
        self.tray.show()
        self.tray.setToolTip("Network Nerdathron 3001™ is watching 👁")

    def _start_monitor(self):
        self.monitor = Monitor(self.engine)
        self.monitor.new_conn.connect(self._on_new_conn)
        self.monitor.all_conns.connect(self.live_panel.update)
        self.monitor.start()
        self._monitoring = True

    def _toggle_monitor(self):
        if self._monitoring:
            self.monitor.paused = True
            self._monitoring = False
            self.pause_btn.setText("▶ RESUME")
            self.monitor_lbl.setText("● PAUSED")
            self.monitor_lbl.setStyleSheet(f"color:{C['warn']};font-size:11px;font-weight:bold;background:transparent;")
        else:
            self.monitor.paused = False
            self._monitoring = True
            self.pause_btn.setText("⏸ PAUSE")
            self.monitor_lbl.setText("● MONITORING")
            self.monitor_lbl.setStyleSheet(f"color:{C['lime']};font-size:11px;font-weight:bold;background:transparent;")

    def _reset_seen(self):
        if hasattr(self, 'monitor'):
            self.monitor.seen.clear()
        self.statusBar().showMessage("🔄  Seen-connections cache cleared — all connections will be re-evaluated")

    def _setup_timers(self):
        self._pulse_state = True
        self._pulse = QTimer(self)
        self._pulse.timeout.connect(self._do_pulse)
        self._pulse.start(700)

        self._refresh_timer = QTimer(self)
        self._refresh_timer.timeout.connect(self._auto_refresh)
        self._refresh_timer.start(20000)

    def _do_pulse(self):
        if not self._monitoring:
            return
        self._pulse_state = not self._pulse_state
        c = C['lime'] if self._pulse_state else C['teal']
        self.monitor_lbl.setStyleSheet(f"color:{c};font-size:11px;font-weight:bold;background:transparent;")

    def _auto_refresh(self):
        idx = self.tabs.currentIndex()
        if idx == 0:
            self.rules_panel.refresh()
        elif idx == 2:
            self.stats_panel.refresh()

    def _on_new_conn(self, conn: ConnInfo):
        decision = self.engine.check(conn)

        if decision is not None:
            self.db.log(conn, decision)
            self.log_panel.add_live(conn, decision)
            icon = '✅' if decision == 'allow' else '🚫'
            self.statusBar().showMessage(
                f"{icon} [{decision.upper()}]  {conn.process}  →  {conn.remote_ip}:{conn.remote_port}"
                f"  [{conn.service()}]"
            )
            if decision == 'allow':
                self._allow_count += 1
                self.allow_lbl.setText(f"✅ {self._allow_count}")
            else:
                self._deny_count += 1
                self.deny_lbl_hdr.setText(f"🚫 {self._deny_count}")
                self._try_terminate(conn)
            return

        if self._pending:
            self.db.log(conn, 'pending')
            self.log_panel.add_live(conn, 'pending')
            return

        self._pending = True
        self._alert_count += 1
        self.alert_lbl.setText(f"Alerts: {self._alert_count}")

        self.show()
        self.raise_()
        self.activateWindow()

        dlg = AlertDialog(conn, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            action = dlg.result_action or 'allow'
            rule   = dlg.result_rule

            if rule:
                self.engine.add(rule)
                self.rules_panel.refresh()

            self.db.log(conn, action)
            self.log_panel.add_live(conn, action)

            icon = '✅' if action == 'allow' else '🚫'
            self.statusBar().showMessage(
                f"{icon} USER DECISION [{action.upper()}]  {conn.process}"
                f"  →  {conn.remote_ip}:{conn.remote_port}"
            )
            if action == 'allow':
                self._allow_count += 1
                self.allow_lbl.setText(f"✅ {self._allow_count}")
            else:
                self._deny_count += 1
                self.deny_lbl_hdr.setText(f"🚫 {self._deny_count}")
                self._try_terminate(conn)

            self.tray.showMessage(
                "Nerdathron 3001™",
                f"{icon} {action.upper()}: {conn.process} → {conn.remote_ip}:{conn.remote_port}",
                QSystemTrayIcon.MessageIcon.Information, 2000
            )

        self._pending = False

    def _try_terminate(self, conn: ConnInfo):
        try:
            proc = psutil.Process(conn.pid)
            for c in proc.net_connections():
                if c.raddr and c.raddr.ip == conn.remote_ip and c.raddr.port == conn.remote_port:
                    proc.terminate()
                    self.statusBar().showMessage(
                        f"🔪  Terminated process {conn.process} (pid {conn.pid}) — connection denied"
                    )
                    break
        except Exception:
            pass

    def closeEvent(self, event):
        event.ignore()
        self.hide()
        self.tray.showMessage(
            "Network Nerdathron 3001™",
            "Still watching your network from the tray, baby! 👁🌈",
            QSystemTrayIcon.MessageIcon.Information, 3000
        )


# ══════════════════════════════════════════════════════════════════════
#  🚀  LAUNCH PAD
# ══════════════════════════════════════════════════════════════════════

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Network Nerdathron 3001")
    app.setApplicationVersion("2.0.0")
    app.setStyle("Fusion")

    is_root = (os.geteuid() == 0)
    if not is_root:
        msg = QMessageBox()
        msg.setStyleSheet(STYLESHEET)
        msg.setWindowTitle("⚠  Privilege Check")
        msg.setIcon(QMessageBox.Icon.Warning)
        msg.setText(
            f"<b style='color:#FFE000;font-size:14px;'>Network Nerdathron 3001™</b><br><br>"
            "<b>Running WITHOUT root privileges.</b><br><br>"
            "✅  Connection MONITORING will work fine.<br>"
            "⚠️  Connection TERMINATION requires:<br>"
            "&nbsp;&nbsp;&nbsp;<code>sudo python3 nerdathron3001.py</code><br><br>"
            "Continue in monitoring-only mode?"
        )
        msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if msg.exec() == QMessageBox.StandardButton.No:
            sys.exit(0)

    window = Nerdathron3001()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
